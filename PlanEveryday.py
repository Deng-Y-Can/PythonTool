import os
import tkinter as tk
from tkinter import filedialog
import openpyxl
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime, timedelta
from tkcalendar import Calendar
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import time
from datetime import date
import schedule

mainExcelPath=os.path.join(os.getcwd(),"plan.xlsx")
def check_and_create_plan_file(path,file_name = 'plan.xlsx'):
    full_path = os.path.join(path, file_name)
    if not os.path.exists(full_path):
        workbook = openpyxl.Workbook()
        workbook.save(full_path)
        return False
    else:
        return True

#指定一个路径创建excel文件
def create_plan_file(file_name = 'plan.xlsx'):
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askdirectory()
    if path:
        full_path = os.path.join(path, file_name)
        if not os.path.exists(full_path):
            workbook = openpyxl.Workbook()
            workbook.save(full_path)
            show_message(message=f"在路径 {path} 下成功创建 plan.xlsx 文件。")
        else:
            show_message(message=f"路径 {path} 下已存在 plan.xlsx 文件。")
    else:
        show_message(message="未选择路径。")

#指定一个路径创建excel文件,自主命名
def create_custom_excel_file():
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askdirectory()
    if path:
        excel_name = tk.simpledialog.askstring("输入文件名", "请输入要创建的 Excel 文件名称：")
        if excel_name:
            full_path = os.path.join(path, f"{excel_name}.xlsx")
            if not os.path.exists(full_path):
                workbook = openpyxl.Workbook()
                workbook.save(full_path)
                show_message(message=f"在路径 {path} 下成功创建 {excel_name}.xlsx 文件。")
            else:
                show_message(message=f"路径 {path} 下已存在名为 {excel_name}.xlsx 的文件。")
        else:
            show_message(message="未输入文件名。")
    else:
        show_message(message="未选择路径。")

#create_custom_excel_file()

def show_message(title="温馨提示",message=""):
    messagebox.showinfo(title, message)



def show_confirm_dialog(title="温馨提示",message=""):
    response = messagebox.askokcancel(title,message)
    if response:
        return True
    else:
        return False



def open_and_create_sheet(excel_file_path,sheet="parameter"):
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # 创建名为 sheet 的工作表
    if sheet not in workbook.sheetnames:
        workbook.create_sheet(sheet)

    workbook.save(excel_file_path)
    return workbook


def add_column_to_parameter_sheet(excel_file_path,sheet,columnName):
    workbook = openpyxl.load_workbook(excel_file_path)
    if sheet in workbook.sheetnames:
        sheet = workbook[sheet]
        sheet.cell(row=1, column=sheet.max_column + 1, value=columnName)
        workbook.save(excel_file_path)
    else:
        show_message("工作表'parameter'不存在。")

def add_row_to_parameter_sheet(excel_file_path,sheetName,column,value):
    workbook = openpyxl.load_workbook(excel_file_path)
    if sheetName in workbook.sheetnames:
        sheet = workbook[sheetName]
        new_row_number = sheet.max_row + 1
        for i in range(len(column)):
            sheet.cell(row=new_row_number, column=get_column_number(sheet, column[i]), value=value[i])
        workbook.save(excel_file_path)
    else:
        show_message("工作表'parameter'不存在。")

def find_rows_in_excel(file_path,sheetName,columnName,values):
    workbook = openpyxl.load_workbook(file_path)
    if sheetName in workbook.sheetnames:
        sheet = workbook[sheetName]
        createdate_column = None
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                if cell.value == columnName:
                    createdate_column = cell.column
                    break
            if createdate_column:
                break

        matching_rows = []
        for row in sheet.iter_rows(min_row=2):
            if row[createdate_column - 1].value==values:
                matching_rows.append([cell.value for cell in row])
        return matching_rows
    else:
        return None

def get_column_number(sheet, column_name):
    for col_num, cell in enumerate(sheet[1], start=1):
        if cell.value == column_name:
            return col_num
    return None

def get_email(title="输入邮箱",message="请输入邮箱："):
    def on_ok():
        global email
        email = entry.get()
        dialog.destroy()

    def on_cancel():
        global email
        email = None
        dialog.destroy()

    dialog = tk.Toplevel()
    dialog.title(title)
    dialog.geometry("200x150")

    label = tk.Label(dialog, text=message)
    label.pack()

    entry = tk.Entry(dialog)
    entry.pack()

    ok_button = tk.Button(dialog, text="确定", command=on_ok)
    ok_button.pack(side=tk.RIGHT)
    cancel_button = tk.Button(dialog, text="取消", command=on_cancel)
    cancel_button.pack(side=tk.RIGHT)

    dialog.wait_window()
    return email

def show_dialog():

    def on_ok():
        global selected_date
        global task
        selected_date = calendar.get_date()
        #task_entry.focus_set()
        task = task_entry.get()
        dialog.destroy()
        #return selected_date, task

    def on_cancel():
        global selected_date
        global task
        selected_date = None
        task = None
        dialog.destroy()
        #return None, None

    dialog = tk.Toplevel()
    dialog.title("计划设置")
    dialog.geometry("400x300")

    # 选择计划日期的控件
    date_label = tk.Label(dialog, text="选择计划日期:")
    date_label.pack()
    calendar = Calendar(dialog, selectmode='day', date_pattern='yyyy-mm-dd')
    calendar.pack()

    # 任务输入框
    task_label = tk.Label(dialog, text="任务:")
    task_label.pack()
    task_entry = tk.Entry(dialog)
    task_entry.pack()
    task_entry.focus_set()

    ok_button = tk.Button(dialog, text="确定", command=on_ok)
    ok_button.pack(side=tk.RIGHT)
    cancel_button = tk.Button(dialog, text="取消", command=on_cancel)
    cancel_button.pack(side=tk.RIGHT)

    dialog.wait_window()
    return selected_date, task


def show_dialog_plan():
    dialog = tk.Toplevel()
    dialog.title("计划设置")
    dialog.geometry("400x300")

    # 计算十年后的日期
    ten_years_from_now = datetime.now() + timedelta(days=365 * 1)

    # 选择计划日期的控件
    date_label = tk.Label(dialog, text="选择计划日期:")
    date_label.pack()
    date_combo = ttk.Combobox(dialog)
    current_date = datetime.now()
    for i in range(0, 365 * 1):
        next_date = current_date + timedelta(days=i)
        date_combo['values'] = (*date_combo['values'], next_date.strftime('%Y-%m-%d'))
    date_combo.pack()

    # 任务输入框
    task_label = tk.Label(dialog, text="任务:")
    task_label.pack()
    task_entry = tk.Entry(dialog)
    task_entry.pack()

    def on_ok():
        selected_date = date_combo.get()
        task = task_entry.get()
        return {'date': selected_date, 'task': task}
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    ok_button = tk.Button(dialog, text="确定", command=on_ok)
    ok_button.pack(side=tk.RIGHT)
    cancel_button = tk.Button(dialog, text="取消", command=on_cancel)
    cancel_button.pack(side=tk.RIGHT)

def initialization():
    open_and_create_sheet(mainExcelPath,"plan" )
    open_and_create_sheet(mainExcelPath)
    useremail=get_email()
    add_column_to_parameter_sheet(mainExcelPath,"parameter","key")
    add_column_to_parameter_sheet(mainExcelPath,"parameter", "values")
    add_column_to_parameter_sheet(mainExcelPath, "plan", "createDate")
    add_column_to_parameter_sheet(mainExcelPath, "plan", "completeDtime")
    add_column_to_parameter_sheet(mainExcelPath, "plan", "task")
    add_column_to_parameter_sheet(mainExcelPath, "plan", "isComplete")
    add_row_to_parameter_sheet(mainExcelPath,"parameter",["key","values"],["email",useremail])
    show_message(f"初始化完成！")

def task():
    if show_confirm_dialog("是否创建学习计划？"):
        selected_date,task=show_dialog()
        if selected_date !=None:
            add_row_to_parameter_sheet(mainExcelPath, "plan", ["createDate", "completeDtime","task","isComplete"], [datetime.now(), selected_date,task,"False"])
            #add_row_to_parameter_sheet(mainExcelPath, "plan", ["createDate", "completeDtime","task","isComplete"], [datetime.now(), result["date"],result["task"],"False"])
            show_message(f"学习计划已创建，加油！")

def send_email2(to_email,content):
    # 邮件服务器地址和端口（以 QQ 邮箱为例）
    smtp_server = 'smtp.qq.com'
    smtp_port = 587  #587  465
    smtp_port_ssl=465
    # 发件人邮箱地址
    sender_email = '3489819802@qq.com'
    sender_password = 'yrnyhuklvzqtcjhh'

    # 邮件内容
    subject = '学习提醒'
    # content = f"""您好：
    # 是时候学习了哦！
    # 你的小助手
    # {datetime.now().date()}"""
    message = MIMEText(content, 'plain', 'utf-8')
    message['From'] = Header(sender_email)
    message['To'] = Header(to_email)
    message['Subject'] = Header(subject)

    max_retries = 3
    retry_delay = 5

    for attempt in range(max_retries):
        try:
            # 连接服务器
            server = smtplib.SMTP(smtp_server, smtp_port)
            #server = smtplib.SMTP_SSL(smtp_server, smtp_port_ssl)
            server.starttls()
            # 登录发件人邮箱
            server.login(sender_email, sender_password)
            # 发送邮件
            server.sendmail(sender_email, to_email, message.as_string())
            print(f"邮件已发送至 {to_email}")
            return
        except Exception as e:
            print(f"发送邮件时出现错误，尝试 {attempt + 1}/{max_retries} 次：{e}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
    print("多次尝试后仍无法发送邮件。")

def timecheck():
    result=find_rows_in_excel(mainExcelPath, "plan", "completeDtime", date.today().strftime('%Y-%m-%d'))
    for item in result:
        text=f"""亲爱的：
    是时候学习了哦！
    任务：{item[3]}
    你的小助手
    {datetime.now().date()}"""
        send_email2("3263554644@qq.com", text)
    show_message(f"测试完成")


def is_today(string_date):
    today = datetime.now().date()
    try:
        parsed_date = datetime.strptime(string_date, '%Y-%m-%d').date()
        return parsed_date == today
    except ValueError:
        return False

#主程序开始
path = os.getcwd()
result = check_and_create_plan_file(path)
if not result:
    initialization()
show_message(f"您好，我滴宝贝，欢迎使用学习监督小能手！")
task()

# 设置每天 8 点执行方法
schedule.every().day.at("11:43").do(timecheck())

while True:
    schedule.run_pending()
    time.sleep(1)

#task()
#send_email2("3263554644@qq.com")
#timecheck()







